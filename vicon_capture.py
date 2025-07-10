# Written By: Brandon Johns
# Date Version Created: 2022-11-28
# Date Last Edited: 2022-11-28
# Status: functional

### PURPOSE ###
# Example for using Brandon's interface to the 'Vicon DataStream SDK'
# Output object location & orientation to excel

### NOTES ###
# Due to use of the 'multiprocessing' module,
# for some reason, any script that uses this module must be wrapped in
#   if __name__ == '__main__':
# Otherwise there will be some weird error about needing to call
#   multiprocessing.freeze_support()
# See: https://stackoverflow.com/a/24374798


import VDSInterface
import time
import numpy as np
import openpyxl
from datetime import datetime
import threading
import os

# Utility for openpyxl
#   In openpyxl, the first cell is at column=1,row=1
def writeRow(sheet, row, data):
    idx = 0 # fight me
    if isinstance(data, np.ndarray):
        for value in data.flat:
            sheet.cell(column=idx+1, row=row, value=value)
            idx += 1
    elif isinstance(data, list) or isinstance(data, tuple):
        for value in data:
            sheet.cell(column=idx+1, row=row, value=value)
            idx += 1
    else:
        # Single value, not an array
        sheet.cell(column=idx+1, row=row, value=data)


if __name__ == "__main__":
    # Program configuration
    hostName = '192.168.11.3:801' # IP address of the computer running Vicon Tracker
    lightWeightMode = False
    excelFilename = 'rawData_preso_test_' + datetime.today().strftime('%Y-%m-%d_%H-%M-%S') + '.xlsx'

    # Create Excel file
    # with sheets for position and orientation of each object
    workbook = openpyxl.Workbook()
    workbook.active.title = 'FrameInfo'
    workbook.create_sheet('Glass_P')
    workbook.create_sheet('Glass_R')
    # Header row
    headerFrameInfo = ['Number','Time']
    headerP = ['x', 'y', 'z']
    headerR = ['qw', 'qx', 'qy', 'qz']
    writeRow(workbook['FrameInfo'], 1, headerFrameInfo)
    writeRow(workbook['Glass_P'], 1, headerP)
    writeRow(workbook['Glass_R'], 1, headerR)

    # Instance the interface
    vdsi = VDSInterface.Interface()
    vdsi.Connect(hostName, lightWeightMode)

    timeStart = time.time()
    excelRow = 2 # First row has the header. Data starts at row 2
    # Collect data
    
    stop_flag = False
    def wait_for_input():
        global stop_flag
        print('Starting in 3 seconds...')
        input('\033[36mPress ENTRE to stop...\033[0m\n')
        stop_flag = True
    threading.Thread(target=wait_for_input, daemon=True).start()
    
    try:
        time.sleep(3)
        while not stop_flag:
            # Get next data frame
            frame = vdsi.GetFrame_GetUnread()
            Glass = frame.GetByName('hk_test')

            # If the object is occluded, the row will be blank
            print('Frame:', frame.FrameNumber())
            position = ['nan' if np.sum(Glass.P())==np.nan else Glass.P()]
            rotation = ['nan' if np.sum(Glass.quat_wxyz())==np.nan else Glass.quat_wxyz()]
            print('Glass.P()', Glass.P())
            print('Glass.quat_wxyz()', Glass.quat_wxyz())
            
            writeRow(workbook['FrameInfo'], excelRow, [frame.FrameNumber(), frame.FrameTime_seconds()-timeStart])
            writeRow(workbook['Glass_P'], excelRow, Glass.P())
            writeRow(workbook['Glass_R'], excelRow, Glass.quat_wxyz())
            print(f"Wrote row {excelRow} to Excel")

            excelRow += 1
        
    finally:
        print('\033[36mCleaning up...\033[0m')
        time.sleep(1)
        vdsi.Disconnect()
        workbook.save(excelFilename)
        print(f"Excel file saved at: {os.path.abspath(excelFilename)}")